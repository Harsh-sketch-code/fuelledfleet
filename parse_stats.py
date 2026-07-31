"""
Parse TitanGPS Driver Statistics Report XLSX files (per-driver, per-week)
into structured JSON, then merge with the Fleet Report PDF parses to produce
a single combined weeks.json the dashboard reads.

Usage:
    python3 parse_stats.py
"""
import json, os, re, glob
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
STATS_DIR = os.path.join(BASE, "driver statistics report")
WEEKS_DIR = os.path.join(BASE, "weekly reports")
OUT_FILE  = os.path.join(BASE, "weeks.json")

NAME_MAP = {
    "Dylan":   "d1",
    "Will":    "d2", "Willem": "d2",
    "Jackson": "d3",
    "Austin":  "d4",
    "Paddy":   "d5", "Patrick": "d5", "Raj": "d5",
    "Harsh":   "d6",
}

ROSTER = [
    {"id":"d1", "name":"Dylan",   "truck":"023 - Canyon AT4X"},
    {"id":"d2", "name":"Willem",  "truck":"022 - Canyon AT4X"},
    {"id":"d3", "name":"Jackson", "truck":"021 - Canyon AT4X"},
    {"id":"d4", "name":"Austin",  "truck":"004 - Ford F150"},
    # Truck 020 (Canyon AT4X, formerly 005 Ford F150): Raj drove through April 2026 only;
    # Paddy takes over starting May 1, 2026 and moved to the new Canyon in August 2026.
    {"id":"d5", "name":"Paddy",   "truck":"020 - Canyon AT4X",
     "name_history": [
        {"from_iso": "2000-01-01", "name": "Raj"},
        {"from_iso": "2026-05-01", "name": "Paddy"},
     ]},
    {"id":"d6", "name":"Harsh",   "truck":"019 - Canyon AT4X"},
]

# Column header (case-insensitive contains) -> our key
COLS = {
    "minutes analyzed":              "minutes",
    "driver score":                  "score_xlsx",
    "average following distance(sec)":"avg_fd_sec",
    "sign violations":               "sign_violations",
    "traffic light violation":       "traffic_lights",
    "hard braking":                  "hard_brake",
    "hard acceleration":             "hard_accel",
    "following distance":            "following_close",
    "speeding violations":           "speeding",
    "driver star":                   "stars",
}
SUFFIX_SKIP = "gz impact"  # ignore the *_GZ_Impact columns

def to_int(v, default=0):
    try:
        if v is None or v == "" or v == "NA":
            return default
        return int(float(v))
    except Exception:
        return default

def to_float(v, default=0.0):
    try:
        if v is None or v == "" or v == "NA":
            return default
        return float(v)
    except Exception:
        return default

def header_index(row):
    """Return {our_key: column_index} based on the header row."""
    idx = {}
    for i, cell in enumerate(row):
        if cell is None: continue
        h = str(cell).strip().lower()
        # skip the GZ Impact columns (they live to the right of the actual count)
        if h.endswith(SUFFIX_SKIP):
            continue
        for needle, key in COLS.items():
            if h == needle or h.startswith(needle):
                # First match wins (so "Following Distance" is matched, not
                # "Average Following Distance(Sec)" — handled by COLS keys)
                if key not in idx:
                    idx[key] = i
                break
    return idx

def parse_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # Find period from row containing "Report Generated For"
    period_iso_start = period_iso_end = period_label = None
    for r in rows[:6]:
        for cell in r:
            if cell and "Report Generated For" in str(cell):
                m = re.search(r"\((\d{1,2}) (\w+)\) [\d:]+ [AP]M\s*\([A-Z]+\)\s*to\s*\w+\s*\((\d{1,2}) (\w+)\)", str(cell))
                if m:
                    sd, sm, ed, em = m.groups()
                    months = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
                              "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
                    sm_n, em_n = months.get(sm[:3]), months.get(em[:3])
                    fname = os.path.basename(path)
                    yr_m = re.search(r"(20\d{2})", fname)
                    yr = int(yr_m.group(1)) if yr_m else 2026
                    if sm_n and em_n:
                        period_iso_start = f"{yr}-{sm_n:02d}-{int(sd):02d}"
                        end_yr = yr + (1 if em_n < sm_n else 0)
                        # XLSX end date is exclusive (Mon to Mon = 7 days),
                        # subtract 1 day to match weekly Fleet Report convention (Mon-Sun)
                        from datetime import date, timedelta
                        ed_date = date(end_yr, em_n, int(ed)) - timedelta(days=1)
                        period_iso_end = ed_date.isoformat()
                        period_label = f"{sm} {int(sd)}-{ed_date.strftime('%b')} {ed_date.day}, {yr}"
                break
        if period_iso_start:
            break

    # Find the header row (the one with "Driver Name" in col 0)
    header_row_idx = None
    for i, r in enumerate(rows):
        if r and r[0] and str(r[0]).strip().lower() == "driver name":
            header_row_idx = i
            break
    if header_row_idx is None:
        raise RuntimeError(f"Could not find header row in {path}")

    cmap = header_index(rows[header_row_idx])

    drivers = []
    for r in rows[header_row_idx+1:]:
        if not r or not r[0]: continue
        nm = str(r[0]).strip()
        if nm.lower() in ("total", "unknown driver", ""):
            continue
        if nm not in NAME_MAP:
            continue  # unknown driver
        rec = {
            "name": nm,
            "driver_id": NAME_MAP[nm],
            "minutes":          to_int(r[cmap["minutes"]]) if "minutes" in cmap else 0,
            # Score 0 in TitanGPS means "no real driving measured" (typically a brief
            # vehicle ping under a few minutes). Normalize to None.
            "score":            (lambda v: v if v else None)(to_int(r[cmap["score_xlsx"]])) if "score_xlsx" in cmap else None,
            "avg_fd_sec":       to_float(r[cmap["avg_fd_sec"]]) if "avg_fd_sec" in cmap else None,
            "sign_violations":  to_int(r[cmap["sign_violations"]]) if "sign_violations" in cmap else 0,
            "traffic_lights":   to_int(r[cmap["traffic_lights"]]) if "traffic_lights" in cmap else 0,
            "hard_brake":       to_int(r[cmap["hard_brake"]]) if "hard_brake" in cmap else 0,
            "hard_accel":       to_int(r[cmap["hard_accel"]]) if "hard_accel" in cmap else 0,
            "following_close":  to_int(r[cmap["following_close"]]) if "following_close" in cmap else 0,
            "speeding":         to_int(r[cmap["speeding"]]) if "speeding" in cmap else 0,
            "stars":            to_int(r[cmap["stars"]]) if "stars" in cmap else 0,
        }
        # Skip drivers with 0 minutes (Austin in early weeks shows up but with no data)
        if rec["minutes"] == 0 and rec["score"] in (None, 0):
            continue
        drivers.append(rec)

    return {
        "source_file": os.path.basename(path),
        "period": {"start_iso": period_iso_start, "end_iso": period_iso_end, "label": period_label},
        "drivers": drivers,
    }

MONTH_NAME_TO_NUM = {"january":1,"february":2,"march":3,"april":4,"may":5,"june":6,
                     "july":7,"august":8,"september":9,"october":10,"november":11,"december":12,
                     "jan":1,"feb":2,"mar":3,"apr":4,"jun":6,"jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}

def parse_monthly_stats_xlsx(path):
    """
    Parse a monthly Driver Statistics Report XLSX (one row per driver for the whole month).
    Returns dict keyed by driver_id with TitanGPS-canonical monthly totals (minutes, stars,
    score, infractions, avg_fd_sec). This is more accurate than summing weekly XLSX rows
    because TitanGPS sometimes recomputes events server-side after the weekly snapshot.
    Returns None if the file can't be parsed into a month.
    """
    fname = os.path.basename(path)
    # Pull month name + year from filename, e.g. "Driver Statistics Report( May 2026 ).xlsx"
    m = re.search(r"\(\s*(\w+)\s+(20\d{2})\s*\)", fname)
    if not m:
        return None
    mname = m.group(1).strip().lower()
    yr = int(m.group(2))
    mo = MONTH_NAME_TO_NUM.get(mname)
    if not mo:
        return None

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # Find header row
    header_row_idx = None
    for i, r in enumerate(rows):
        if r and r[0] and str(r[0]).strip().lower() == "driver name":
            header_row_idx = i
            break
    if header_row_idx is None:
        return None

    cmap = header_index(rows[header_row_idx])

    drivers = {}
    for r in rows[header_row_idx+1:]:
        if not r or not r[0]: continue
        nm = str(r[0]).strip()
        if nm.lower() in ("total", "unknown driver", ""):
            continue
        if nm not in NAME_MAP:
            continue
        did = NAME_MAP[nm]
        rec = {
            "name": nm,
            "minutes":          to_int(r[cmap["minutes"]]) if "minutes" in cmap else 0,
            "score":            to_int(r[cmap["score_xlsx"]], default=None) if "score_xlsx" in cmap else None,
            "avg_fd_sec":       to_float(r[cmap["avg_fd_sec"]]) if "avg_fd_sec" in cmap else None,
            "sign_violations":  to_int(r[cmap["sign_violations"]]) if "sign_violations" in cmap else 0,
            "traffic_lights":   to_int(r[cmap["traffic_lights"]]) if "traffic_lights" in cmap else 0,
            "hard_brake":       to_int(r[cmap["hard_brake"]]) if "hard_brake" in cmap else 0,
            "hard_accel":       to_int(r[cmap["hard_accel"]]) if "hard_accel" in cmap else 0,
            "following_close":  to_int(r[cmap["following_close"]]) if "following_close" in cmap else 0,
            "speeding":         to_int(r[cmap["speeding"]]) if "speeding" in cmap else 0,
            "stars":            to_int(r[cmap["stars"]]) if "stars" in cmap else 0,
        }
        # If score is 0 and minutes is 0, it's a stub row (driver didn't drive that month)
        if rec["minutes"] == 0 and rec["score"] in (None, 0):
            continue
        # If multiple aliases map to same did (e.g. Raj+Paddy = d5), keep the one with data
        if did in drivers and rec["minutes"] <= drivers[did]["minutes"]:
            continue
        drivers[did] = rec

    return {"year": yr, "month": mo, "source_file": fname, "drivers": drivers}

def load_monthly_stats_xlsx_index():
    """Returns {(year,month): parsed_dict}"""
    monthly_dir = os.path.join(BASE, "monthly reports")
    out = {}
    if not os.path.isdir(monthly_dir):
        return out
    for p in sorted(glob.glob(os.path.join(monthly_dir, "*.xlsx"))):
        parsed = parse_monthly_stats_xlsx(p)
        if parsed:
            out[(parsed["year"], parsed["month"])] = parsed
    return out

def load_pdf_index():
    p = os.path.join(WEEKS_DIR, "_parsed_index.json")
    if not os.path.exists(p):
        return {}
    with open(p) as f:
        idx = json.load(f)
    by_start = {}
    for r in idx.get("reports", []):
        by_start[r["period"]["start_iso"]] = r
    return by_start

def merge():
    pdfs = load_pdf_index()
    weeks = []
    for x in sorted(glob.glob(os.path.join(STATS_DIR, "*.xlsx"))):
        s = parse_xlsx(x)
        start = s["period"]["start_iso"]
        pdf = pdfs.get(start, {})
        # Build per-driver merged record
        per_driver = {}
        for d in s["drivers"]:
            per_driver[d["driver_id"]] = {
                "name": d["name"],
                "score": d["score"],
                "minutes": d["minutes"],
                "avg_fd_sec": d["avg_fd_sec"],
                "stars": d["stars"],
                "infractions": {
                    "sign_violations": d["sign_violations"],
                    "following_close": d["following_close"],
                    "hard_accel":      d["hard_accel"],
                    "hard_brake":      d["hard_brake"],
                    "speeding":        d["speeding"],
                    "traffic_lights":  d["traffic_lights"],
                },
            }
        # Override scores with PDF values where available (PDF = portal-canonical)
        if pdf:
            for p in pdf.get("driver_performance", []):
                did = p["driver_id"]
                if did in per_driver:
                    per_driver[did]["score"] = p["score"]
                    per_driver[did]["delta"] = p["delta"]
                else:
                    # Driver appears in PDF but not in XLSX — keep them anyway
                    per_driver[did] = {
                        "name": p["name"], "score": p["score"], "delta": p["delta"],
                        "minutes": 0, "avg_fd_sec": None, "stars": 0,
                        "infractions": {"sign_violations":0,"following_close":0,"hard_accel":0,
                                        "hard_brake":0,"speeding":0,"traffic_lights":0},
                    }
        weeks.append({
            "id": "w_" + start,
            "label": s["period"]["label"],
            "start_iso": start,
            "end_iso": s["period"]["end_iso"],
            "fleet_average_score": pdf.get("fleet_average_score") if pdf else None,
            "drivers_with_score": pdf.get("drivers_with_greenzone_score") if pdf else None,
            "drivers": per_driver,
            "sources": {
                "stats_xlsx": s["source_file"],
                "fleet_pdf":  pdf.get("source_file") if pdf else None,
            },
        })
    # Dedupe by period start_iso (two XLSX filenames that map to the same week
    # — e.g. "( 11 May ... )" vs "( 11 May ...)" — both get parsed, keep last).
    by_start = {}
    for w in weeks:
        by_start[w["start_iso"]] = w
    weeks = list(by_start.values())
    weeks.sort(key=lambda w: w["start_iso"])

    # Mark any week whose end date is today or in the future as "in progress"
    # — TitanGPS will refresh the canonical numbers the following Monday.
    from datetime import date
    today_iso = date.today().isoformat()
    for w in weeks:
        w["is_partial"] = (w.get("end_iso") or "0000-00-00") >= today_iso
        if w["is_partial"]:
            w["partial_through_iso"] = today_iso

    # ===== Build monthly aggregates =====
    # A month appears as soon as any week within it has data (XLSX or PDF).
    # If a Monthly Fleet Report PDF is present, its scores override (TitanGPS-canonical).
    # Otherwise scores are computed from weekly data (weighted by driving minutes).
    from datetime import date, timedelta

    # Load any monthly Fleet Report JSONs available
    monthly_dir = os.path.join(BASE, "monthly reports")
    monthly_pdfs_by_ym = {}
    if os.path.isdir(monthly_dir):
        idx_path = os.path.join(monthly_dir, "_parsed_index.json")
        if os.path.exists(idx_path):
            with open(idx_path) as f:
                for mp in json.load(f).get("reports", []):
                    sd = date.fromisoformat(mp["period"]["start_iso"])
                    monthly_pdfs_by_ym[(sd.year, sd.month)] = mp

    # Load any monthly Driver Statistics XLSX files (TitanGPS-canonical per-driver totals)
    monthly_stats_by_ym = load_monthly_stats_xlsx_index()

    # Discover all (year, month) pairs that have at least one week of data
    months_with_data = sorted({
        (date.fromisoformat(w["start_iso"]).year,
         date.fromisoformat(w["start_iso"]).month)
        for w in weeks
    })
    # Also include months that have a PDF or monthly XLSX but no weekly data (defensive)
    for ym in list(monthly_pdfs_by_ym.keys()) + list(monthly_stats_by_ym.keys()):
        if ym not in months_with_data:
            months_with_data.append(ym)
    months_with_data = sorted(set(months_with_data))

    months = []
    for (yr, mo) in months_with_data:
        first = date(yr, mo, 1)
        if mo == 12:
            last = date(yr, 12, 31)
        else:
            last = date(yr, mo + 1, 1) - timedelta(days=1)
        mp = monthly_pdfs_by_ym.get((yr, mo))

        per_driver = {r["id"]: {
            "name": r["name"], "score": None, "delta": None,
            "minutes": 0, "stars": 0, "avg_fd_sec": None,
            "active_weeks": 0,   # how many weeks in this month had minutes > 0 for this driver
            "infractions": {"sign_violations":0,"following_close":0,"hard_accel":0,
                            "hard_brake":0,"speeding":0,"traffic_lights":0},
            "_fd_samples": [],
            "_score_samples": [],   # (score, minutes) for weighted average
        } for r in ROSTER}

        # Sum infractions / stars / minutes from weekly XLSX rows in this month
        weeks_in_month = []
        for w in weeks:
            wsd = date.fromisoformat(w["start_iso"])
            if wsd.year == yr and wsd.month == mo:
                weeks_in_month.append(w)
                for did, dd in w["drivers"].items():
                    if did not in per_driver: continue
                    pd = per_driver[did]
                    mins  = dd.get("minutes", 0) or 0
                    score = dd.get("score")
                    pd["minutes"] += mins
                    pd["stars"]   += dd.get("stars", 0) or 0
                    if mins > 0:
                        pd["active_weeks"] += 1
                    if dd.get("avg_fd_sec"):
                        pd["_fd_samples"].append(dd["avg_fd_sec"])
                    if score is not None and mins > 0:
                        pd["_score_samples"].append((score, mins))
                    for k, v in (dd.get("infractions") or {}).items():
                        pd["infractions"][k] = pd["infractions"].get(k, 0) + (v or 0)

        # If a monthly Driver Statistics XLSX is present, it is TitanGPS-canonical for
        # the whole month — override the weekly-summed minutes/stars/infractions/avg_fd
        # with its values. (Weekly snapshots can miss late-week events that TitanGPS
        # backfills after the fact.)
        mstats = monthly_stats_by_ym.get((yr, mo))
        if mstats:
            for did, mrec in mstats["drivers"].items():
                if did not in per_driver: continue
                pd = per_driver[did]
                pd["minutes"]    = mrec["minutes"]
                pd["stars"]      = mrec["stars"]
                pd["avg_fd_sec"] = mrec["avg_fd_sec"]
                pd["infractions"] = {
                    "sign_violations": mrec["sign_violations"],
                    "following_close": mrec["following_close"],
                    "hard_accel":      mrec["hard_accel"],
                    "hard_brake":      mrec["hard_brake"],
                    "speeding":        mrec["speeding"],
                    "traffic_lights":  mrec["traffic_lights"],
                }
                # XLSX score is a useful fallback if no PDF is present.
                if mrec["score"] is not None and pd["score"] is None:
                    pd["score"] = mrec["score"]

        # Score: prefer monthly PDF (canonical); fall back to minutes-weighted avg of weekly
        if mp:
            for p in mp.get("driver_performance", []):
                did = p["driver_id"]
                if did in per_driver:
                    per_driver[did]["score"] = p["score"]
                    per_driver[did]["delta"] = p["delta"]

        for did, pd in per_driver.items():
            samples = pd.pop("_score_samples")
            fd_samples = pd.pop("_fd_samples")
            # Compute weighted score average if PDF didn't already provide one
            if pd["score"] is None and samples:
                total_min = sum(m for _, m in samples)
                if total_min > 0:
                    pd["score"] = round(sum(s * m for s, m in samples) / total_min)
            # Only fall back to weekly FD average if monthly XLSX didn't already set it
            if fd_samples and pd["avg_fd_sec"] is None:
                pd["avg_fd_sec"] = round(sum(fd_samples) / len(fd_samples), 2)

        # Fleet average for the month: prefer PDF, else weighted avg of weekly fleet averages
        fleet_avg = mp.get("fleet_average_score") if mp else None
        if not fleet_avg:
            wfa = [(w["fleet_average_score"]["score"], w.get("fleet_average_score") and 1)
                   for w in weeks_in_month if w.get("fleet_average_score")]
            if wfa:
                fleet_avg = {"score": round(sum(s for s, _ in wfa) / len(wfa)), "delta": None}

        # ===== Monthly winner eligibility =====
        # The monthly winner program started in May 2026. From May onward,
        # a driver must log at least 900 minutes in the month to be eligible
        # for the monthly winner spot (small samples shouldn't crown anyone).
        # Before May 2026, everyone ranks normally — no eligibility gate.
        MIN_MONTHLY_MINUTES = 900
        ELIGIBILITY_START = (2026, 5)
        gate_active = (yr, mo) >= ELIGIBILITY_START
        for did, pd in per_driver.items():
            if gate_active:
                pd["eligible"] = pd["minutes"] >= MIN_MONTHLY_MINUTES
            else:
                pd["eligible"] = True   # rule didn't exist yet

        months.append({
            "id": "m_" + first.strftime("%Y-%m"),
            "label": first.strftime("%B %Y"),
            "start_iso": first.isoformat(),
            "end_iso": last.isoformat(),
            "fleet_average_score": fleet_avg,
            "drivers_with_score": (mp or {}).get("drivers_with_greenzone_score"),
            "drivers": per_driver,
            # In progress only if the month isn't finished yet AND there's no monthly
            # snapshot (PDF or XLSX). Once any monthly snapshot lands the month is final.
            "is_provisional": (mp is None and mstats is None) and last.isoformat() >= date.today().isoformat(),
            "sources": {
                "fleet_pdf": mp.get("source_file") if mp else None,
                "monthly_xlsx": mstats["source_file"] if mstats else None,
                "weekly_aggregated": [w["sources"]["stats_xlsx"] for w in weeks_in_month],
            },
        })

    months.sort(key=lambda m: m["start_iso"])
    return {"roster": ROSTER, "weeks": weeks, "months": months}

if __name__ == "__main__":
    out = merge()
    with open(OUT_FILE, "w") as f:
        json.dump(out, f, indent=2)
    print(f"Wrote {OUT_FILE}")
    print(f"  Weeks: {len(out['weeks'])}")
    for w in out["weeks"]:
        sn = sum(1 for d in w["drivers"].values() if d.get("score"))
        st = sum(d.get("stars",0) for d in w["drivers"].values())
        print(f"   - {w['label']:<28}  drivers w/ score: {sn}  stars: {st}")
